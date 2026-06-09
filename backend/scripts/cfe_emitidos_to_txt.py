import csv
import os
import re
import sys
from datetime import datetime
from decimal import Decimal, ROUND_HALF_UP, InvalidOperation

from openpyxl import load_workbook


HEADER = [
    "DIA", "DEBE", "HABER", "CENTRODECOSTOS", "RUC",
    "CONCEPTO", "MONEDA", "TOTAL", "CODIGODEIVA", "IVA", "LIBRO"
]

CUENTA_VENTAS = "4102"
CUENTA_IVA = "21332"
CUENTA_CAJA = "11111"
MONEDA = "0"
LIBRO = "V"
ESPACIO = " "


def texto(valor):
    if valor is None:
        return ""
    return str(valor).strip()


def fmt_importe(valor):
    if valor is None or valor == "":
        valor = 0
    try:
        dec = Decimal(str(valor))
    except (InvalidOperation, ValueError):
        dec = Decimal("0")
    return str(dec.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP))


def obtener_dia(fecha):
    if isinstance(fecha, datetime):
        return str(fecha.day)
    s = texto(fecha)
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return str(datetime.strptime(s[:10], fmt).day)
        except ValueError:
            pass
    return ESPACIO


def solo_digitos(valor):
    return "".join(ch for ch in texto(valor) if ch.isdigit())


def obtener_ruc(valor_receptor):
    digitos = solo_digitos(valor_receptor)
    return digitos if len(digitos) == 12 else ESPACIO


def limpiar_csv(valor):
    return texto(valor).replace(",", " ").replace("\n", " ").replace("\r", " ")


def buscar_fila_encabezado(ws):
    for row in ws.iter_rows():
        valores = [texto(c.value) for c in row]
        if "Número" in valores and "Fecha comp." in valores and "Mnt. Total" in valores:
            return row[0].row
    raise Exception("No se encontró la fila de encabezados del reporte CFE")


def obtener_columnas(ws, fila_header):
    columnas = {}
    for cell in ws[fila_header]:
        if cell.value:
            columnas[texto(cell.value)] = cell.column
    return columnas


def normalizar_numero(numero):
    if isinstance(numero, float) and numero.is_integer():
        return str(int(numero))
    return texto(numero)


def normalizar_tipo(tipo):
    t = texto(tipo)
    t_lower = t.lower()

    if "nota" in t_lower and "ticket" in t_lower:
        return "NC-Ticket"
    if "nota" in t_lower and "factura" in t_lower:
        return "NC-Factura"
    if "nc" in t_lower and "ticket" in t_lower:
        return "NC-Ticket"
    if "nc" in t_lower and "factura" in t_lower:
        return "NC-Factura"
    if "ticket" in t_lower:
        return "e-Ticket"
    if "factura" in t_lower:
        return "e-Factura"

    return t


def es_nota_credito(tipo_normalizado):
    return tipo_normalizado in {"NC-Factura", "NC-Ticket"}


def importe_con_signo(valor, tipo_normalizado):
    if valor is None or valor == "":
        dec = Decimal("0")
    else:
        dec = Decimal(str(valor))

    if es_nota_credito(tipo_normalizado):
        return -abs(dec)
    return dec


def crear_fila(dia, debe, haber, centro_costos, ruc, concepto, total):
    # Los vacíos intermedios se dejan como espacio; los campos finales quedan vacíos.
    return [
        dia,
        debe,
        haber,
        centro_costos,
        ruc,
        concepto,
        MONEDA,
        fmt_importe(total),
        "",
        "",
        LIBRO,
    ]


def generar_txt(input_path, output_path):
    wb = load_workbook(input_path, data_only=True)
    ws = wb["Page 1"] if "Page 1" in wb.sheetnames else wb[wb.sheetnames[0]]

    fila_header = buscar_fila_encabezado(ws)
    cols = obtener_columnas(ws, fila_header)

    requeridas = ["Ser.", "Número", "Fecha comp.", "Receptor", "Rz. Social Rec.", "Mnt. Neto", "Total IVA", "Mnt. Total"]
    faltantes = [c for c in requeridas if c not in cols]
    if faltantes:
        raise Exception("Faltan columnas requeridas: " + ", ".join(faltantes))

    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)

    with open(output_path, "w", newline="", encoding="cp1252") as f:
        writer = csv.writer(f, delimiter=",", lineterminator="\n")
        writer.writerow(HEADER)

        tipo_comprobante = ""

        for fila in range(fila_header + 1, ws.max_row + 1):
            # Las filas tipo "Tipo comprobante  : e-Factura" vienen en la primera columna.
            marca_tipo = texto(ws.cell(fila, 1).value) or texto(ws.cell(fila, 2).value)
            if "Tipo comprobante" in marca_tipo:
                tipo_raw = marca_tipo.split(":")[-1].strip()
                tipo_comprobante = normalizar_tipo(tipo_raw)
                continue

            # Saltar filas de totales/subtotales.
            primera = texto(ws.cell(fila, 1).value)
            if primera.lower().startswith("totales"):
                continue

            serie = limpiar_csv(ws.cell(fila, cols["Ser."]).value)
            numero = ws.cell(fila, cols["Número"]).value
            fecha = ws.cell(fila, cols["Fecha comp."]).value
            receptor_doc = ws.cell(fila, cols["Receptor"]).value
            razon_social = ws.cell(fila, cols["Rz. Social Rec."]).value

            if not numero or not fecha or not razon_social:
                continue

            dia = obtener_dia(fecha)
            ruc = obtener_ruc(receptor_doc)
            numero_txt = normalizar_numero(numero)
            concepto = f"{tipo_comprobante}-{serie}-{numero_txt}-{limpiar_csv(razon_social)}"

            neto = importe_con_signo(ws.cell(fila, cols["Mnt. Neto"]).value, tipo_comprobante)
            iva = importe_con_signo(ws.cell(fila, cols["Total IVA"]).value, tipo_comprobante)
            total = importe_con_signo(ws.cell(fila, cols["Mnt. Total"]).value, tipo_comprobante)

            writer.writerow(crear_fila(dia, ESPACIO, CUENTA_VENTAS, ESPACIO, ruc, concepto, neto))
            writer.writerow(crear_fila(dia, ESPACIO, CUENTA_IVA, ESPACIO, ruc, concepto, iva))
            writer.writerow(crear_fila(dia, CUENTA_CAJA, ESPACIO, ESPACIO, ruc, concepto, total))


def main():
    if len(sys.argv) < 3:
        print("Uso: python cfe_emitidos_to_txt.py <input.xlsx> <output.txt>")
        sys.exit(1)

    try:
        generar_txt(sys.argv[1], sys.argv[2])
        print("OK. Archivo creado en:", sys.argv[2])
    except Exception as e:
        print("ERROR:", str(e), file=sys.stderr)
        sys.exit(1)


if __name__ == "__main__":
    main()
